import discord
import asyncio
import openpyxl
import datetime
import bs4
import os
from discord.ext import commands
from urllib.request import urlopen, Request
import urllib.request as req
from urllib.request import URLError
from urllib.request import HTTPError
from urllib.request import Request, urlopen
import urllib
import urllib.request
from urllib.parse import quote
from discord import Client
from bs4 import BeautifulSoup
import time
import warnings
import re

client: Client = discord.Client()


@client.event
async def on_ready():
    print(client.user.id)
    print("ready")
    game = discord.Game("24시간 출장마사지")
    await client.change_presence(status=discord.Status.online, activity=game)


@client.event
async def on_message(message):
    if message.content.startswith('인하좆'):
        channel = message.channel
        await channel.send('씨발 인하봇이야 병신아')

    if message.content.startswith('!사진'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '인하' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="귀요미", description='홀리몰리'.format(msg), color=0xFFE400)
        embed.set_footer(text="뿌우!!")
        embed.set_image(
            url="https://media.discordapp.net/attachments/754846842789036134/765475786454859806/image0.jpg?width=385&height=523")
        await message.channel.send('귀욤짤'.format(msg), embed=embed)

    if message.content.startswith('!조스케'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '사진' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="크킄", description='크크킄'.format(msg), color=0xFFE400)
        embed.set_footer(text="사륜안!")
        embed.set_image(
            url="https://media.discordapp.net/attachments/753874628123951107/766101653207121940/adf.png")
        await message.channel.send('조스케'.format(msg), embed=embed)

    if message.content.startswith('!1반'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '시간표' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="시간표", description='3-1'.format(msg), color=0xFFE400)
        embed.set_footer(text="시간표가 달라질경우 달라진 시간표가 표시되지 않습니다")
        embed.set_image(url="https://media.discordapp.net/attachments/754846842789036137/766090095614492683/unknown.png")
        await message.channel.send('1반 시간표'.format(msg), embed=embed)

    if message.content.startswith('!3반'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '시간표' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="3반 시간표", description='3-3'.format(msg), color=0xFFE400)
        embed.set_footer(text="시간표가 달라질경우 달라진 시간표가 표시되지 않습니다")
        embed.set_image(url="https://media.discordapp.net/attachments/754846842789036137/766090296303550485/unknown.png")
        await message.channel.send('3반 시간표'.format(msg), embed=embed)

    if message.content.startswith('!8반'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '시간표' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="8반 시간표", description='3-8'.format(msg), color=0xFFE400)
        embed.set_footer(text="시간표가 달라질경우 달라진 시간표가 표시되지 않습니다")
        embed.set_image(url="https://media.discordapp.net/attachments/754846842789036137/766090647563141150/unknown.png")
        await message.channel.send('8반 시간표'.format(msg), embed=embed)

    if message.content.startswith('!2반'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '시간표' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="2반 시간표", description='3-2'.format(msg), color=0xFFE400)
        embed.set_footer(text="시간표가 달라질경우 달라진 시간표가 표시되지 않습니다")
        embed.set_image(url="https://media.discordapp.net/attachments/754846842789036137/766090216670625822/unknown.png")
        await message.channel.send('2반 시간표'.format(msg), embed=embed)

    if message.content.startswith('!4반'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '시간표' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="4반 시간표", description='3-4'.format(msg), color=0xFFE400)
        embed.set_footer(text="시간표가 달라질경우 달라진 시간표가 표시되지 않습니다")
        embed.set_image(url="https://media.discordapp.net/attachments/754846842789036137/766090365978542080/unknown.png")
        await message.channel.send('4반 시간표'.format(msg), embed=embed)

    if message.content.startswith('!5반'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '시간표' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="5반 시간표", description='3-5'.format(msg), color=0xFFE400)
        embed.set_footer(text="시간표가 달라질경우 달라진 시간표가 표시되지 않습니다")
        embed.set_image(url="https://media.discordapp.net/attachments/754846842789036137/766090446727413810/unknown.png")
        await message.channel.send('5반 시간표'.format(msg), embed=embed)

    if message.content.startswith('!명령어'):
        embed = discord.Embed(title=f"명령어 목록", description=''.format(), color=0x00E4FF)
        embed.add_field(name="!배워", value="'!배워'로 가르쳐준 다음에 '인하봇' 입력후 가르쳐준 단어를 입력하시면 됩니다.", inline=False)
        embed.add_field(name="!n반", value="'!n반'->'시간표' 입력하시면 시간표가 나옵니다.", inline=False)
        embed.add_field(name="!사진", value="'!사진'->'인하' 입력하시면 귀여운 인하사진이 나옵니다.", inline=False)
        embed.add_field(name="!조스케", value="'!조스케'->'사진' 입력하시면 조스케사진이 나옵니다.", inline=False)
        embed.add_field(name='!롤', value='!롤 닉네임 형식으로 적으면 그 닉네임에대한 정보를 알려줍니다..', inline=False)
        await message.channel.send(embed=embed)

    if message.content.startswith('!도움말'):
        embed = discord.Embed(title=f"명령어 목록", description=''.format(), color=0x00E4FF)
        embed.add_field(name="!배워", value="'!배워'로 가르쳐준 다음에 '인하봇' 입력후 가르쳐준 단어를 입력하시면 됩니다.", inline=False)
        embed.add_field(name="!n반", value="'!n반'->'시간표' 입력하시면 시간표가 나옵니다.", inline=False)
        embed.add_field(name="!사진", value="'!사진'->'인하' 입력하시면 귀여운 인하사진이 나옵니다.", inline=False)
        embed.add_field(name="!조스케", value="'!조스케'->'사진' 입력하시면 조스케사진이 나옵니다.", inline=False)
        embed.add_field(name='!롤', value='!롤 닉네임 형식으로 적으면 그 닉네임에대한 정보를 알려줍니다..', inline=False)
        await message.channel.send(embed=embed)

    if message.content.startswith('!7반'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '시간표' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="7반 시간표", description='3-7'.format(msg), color=0xFFE400)
        embed.set_footer(text="시간표가 달라질경우 달라진 시간표가 표시되지 않습니다")
        embed.set_image(url="https://media.discordapp.net/attachments/754846842789036137/766090579909804032/unknown.png")
        await message.channel.send('7반 시간표'.format(msg), embed=embed)

    if message.content.startswith('!6반'):
        channel = message.channel
        await channel.send('뭐')

        def check(k):
            return k.content == '시간표' and k.channel == channel

        msg = await client.wait_for('message', check=check)
        embed = discord.Embed(title="6반 시간표", description='3-6'.format(msg), color=0xFFE400)
        embed.set_footer(text="시간표가 달라질경우 달라진 시간표가 표시되지 않습니다")
        embed.set_image(url="https://media.discordapp.net/attachments/754846842789036137/766090510125236264/unknown.png")
        await message.channel.send('7반 시간표'.format(msg), embed=embed)

    if message.content.startswith('!배워'):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 1001):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send('가르쳐 주셔서 감사해요 주인님! 인하는 새로운걸 배워서 너무 기뻐요!')
                break
        file.save("기억.xlsx")

    if message.content.startswith('인하봇'):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 1001):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" +str(i)].value)
                break

    if message.content.startswith('인하야'):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 1001):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" +str(i)].value)
                break

    if message.content.startswith("!롤"):
        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)

        url = "http://www.op.gg/summoner/userName=" + enc_location
        html = urllib.request.urlopen(url)

        bsObj = bs4.BeautifulSoup(html, "html.parser")
        rank1 = bsObj.find("div", {"class": "TierRankInfo"})
        rank2 = rank1.find("div", {"class": "TierRank"})
        rank3 = rank2.text
        if rank3 != 'Unranked':
            jumsu1 = rank1.find("div", {"class": "TierInfo"})
            jumsu2 = jumsu1.find("span", {"class": "LeaguePoints"})
            jumsu3 = jumsu2.text
            jumsu4 = jumsu3.strip()

            winlose1 = jumsu1.find("span", {"class": "WinLose"})
            winlose2 = winlose1.find("span", {"class": "wins"})
            winlose2_1 = winlose1.find("span", {"class": "losses"})
            winlose2_2 = winlose1.find("span", {"class": "winratio"})

            winlose2txt = winlose2.text
            winlose2_1txt = winlose2_1.text
            winlose2_2txt = winlose2_2.text


        channel = message.channel
        embed = discord.Embed(
            title='League of Legends 전적',
            description='주인님의 전적이에요...',
            colour=discord.Colour.green()
        )
        if rank3 == 'Unranked':
            embed.add_field(name='당신의 티어', value=rank3, inline=False)
            embed.add_field(name='-당신은 언랭-', value="언랭은 더이상의 정보는 제공하지 않습니다.", inline=False)
            await client.send_message(channel, embed=embed)
        else:
            embed.add_field(name='티어', value=rank3, inline=False)
            embed.add_field(name='LP(점수)', value=jumsu4, inline=False)
            embed.add_field(name='승,패 정보', value=winlose2txt + " " + winlose2_1txt, inline=False)
            embed.add_field(name='승률', value=winlose2_2txt, inline=False)
            embed.set_thumbnail(url="https://media.discordapp.net/attachments/754846842789036137/769109444666589204/710840847c3faa9bf6ef006bdb6d91d0.png?width=446&height=468")
            await message.channel.send(embed=embed)

    if message.content.startswith("!코로나"):
        # 보건복지부 코로나 바이러스 정보사이트"
        covidSite = "http://ncov.mohw.go.kr/index.jsp"
        covidNotice = "http://ncov.mohw.go.kr"
        html = urlopen(covidSite)
        bs = BeautifulSoup(html, 'html.parser')
        latestupdateTime = bs.find('span', {'class': "livedate"}).text.split(',')[0][1:].split('.')
        statisticalNumbers = bs.findAll('span', {'class': 'num'})
        beforedayNumbers = bs.findAll('span', {'class': 'before'})

        # 주요 브리핑 및 뉴스링크
        briefTasks = []
        mainbrief = bs.findAll('a', {'href': re.compile('\/tcmBoardView\.do\?contSeq=[0-9]*')})
        for brf in mainbrief:
            container = []
            container.append(brf.text)
            container.append(covidNotice + brf['href'])
            briefTasks.append(container)
        print(briefTasks)

        # 통계수치
        statNum = []
        # 전일대비 수치
        beforeNum = []
        for num in range(7):
            statNum.append(statisticalNumbers[num].text)
        for num in range(4):
            beforeNum.append(beforedayNumbers[num].text.split('(')[-1].split(')')[0])

        totalPeopletoInt = statNum[0].split(')')[-1].split(',')
        tpInt = ''.join(totalPeopletoInt)
        lethatRate = round((int(statNum[3]) / int(tpInt)) * 100, 2)
        embed = discord.Embed(title="Covid-19 한국 현황", description="", color=0x5CD1E5)
        embed.add_field(name="Latest data refred time",
                        value="해당 자료는 " + latestupdateTime[0] + "월 " + latestupdateTime[1] + "일 " + latestupdateTime[
                            2] + " 자료입니다.", inline=False)
        embed.add_field(name="확진환자(누적)", value=statNum[0].split(')')[-1] + "(" + beforeNum[0] + ")", inline=True)
        embed.add_field(name="완치환자(격리해제)", value=statNum[1] + "(" + beforeNum[1] + ")", inline=True)
        embed.add_field(name="치료중(격리 중)", value=statNum[2] + "(" + beforeNum[2] + ")", inline=True)
        embed.add_field(name="사망", value=statNum[3] + "(" + beforeNum[3] + ")", inline=True)
        embed.add_field(name="누적확진률", value=statNum[6], inline=True)
        embed.add_field(name="치사율", value=str(lethatRate) + " %", inline=True)
        embed.add_field(name="- 최신 브리핑 1 : " + briefTasks[0][0], value="Link : " + briefTasks[0][1], inline=False)
        embed.add_field(name="- 최신 브리핑 2 : " + briefTasks[1][0], value="Link : " + briefTasks[1][1], inline=False)
        embed.set_thumbnail(
            url="https://wikis.krsocsci.org/images/7/79/%EB%8C%80%ED%95%9C%EC%99%95%EA%B5%AD_%ED%83%9C%EA%B7%B9%EA%B8%B0.jpg")
        await message.channel.send("인하봇은 무서워요 ㅠㅠ", embed=embed)

    if message.content.startswith("!급식"):

        url = "http://school.busanedu.net/mojeon-m/main.do"
        html = urllib.request.urlopen(url)
        bs = BeautifulSoup(html, 'html.parser')
        lunch1 = bs.find('div', {'class': 'meal_list_box'})
        lunch2 = lunch1.find('dd', {'class': 'meal_list'})
        lunch3 = lunch2.text
        embed = discord.Embed(title=" 모전중학교 급식 ", description="", color=0x5CD1E5)
        embed.add_field(name=":hamburger: 오늘의급식:hamburger: ", value=lunch3, inline=True)
        embed.set_thumbnail(url="https://media.discordapp.net/attachments/769124476125708288/771302768995467294/4256_shop1_851864.png?width=468&height=468")
        embed.set_footer(text='Service provided by 서재용.', icon_url= 'https://media.discordapp.net/attachments/769229107082297395/771317519499001906/IMG_21092020_115130_256_x_256_.jpg' )
        await message.channel.send(embed=embed)


client.run(os.environ['token']
